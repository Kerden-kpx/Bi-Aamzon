from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import urllib.error
import urllib.request
import uuid

from fastapi import HTTPException, UploadFile

from ..core.config import DEFAULT_BSR_SITE, normalize_site
from ..imports.bsr_importer import import_bsr_data
from ..imports.bsr_monthly_importer import import_bsr_monthly
from ..db import get_connection
from ..repositories import bsr_repo
from ..services.ai_report_prompt import KEEPA_REPORT_SYSTEM_PROMPT, build_keepa_report_user_prompt
from ..services import user_service

_export_daily_jobs_lock = threading.Lock()
_export_daily_jobs: Dict[str, Dict[str, Any]] = {}
_MAX_EXPORT_DAILY_JOBS = 200
_EXPORT_DAILY_TOTAL_URLS = 100
_EXPORT_DAILY_BATCH_SIZE = 10
_BSR_DAILY_REQUIRED_HEADERS = [
    "日期",
    "Buybox价格($)",
    "价格($)",
    "Prime价格($)",
    "Coupon价格($)",
    "Coupon折扣",
    "子体销量",
]

_BSR_DAILY_OPTIONAL_HEADERS = {
    "fba_price": ["FBA价格($)", "FBA价格", "FBA Price($)", "FBA Price"],
    "fbm_price": ["FBM价格($)", "FBM价格", "FBM Price($)", "FBM Price"],
    "strikethrough_price": ["划线价格($)", "划线价格", "List价格($)", "List Price($)"],
    "bsr_rank": ["BSR排名", "BSR Rank"],
    "rating": ["评分", "Rating"],
    "rating_count": ["评分数", "Rating Count", "Review Count"],
    "seller_count": ["卖家数", "Seller Count"],
}
_OPENROUTER_API_BASE = "https://openrouter.ai/api/v1/chat/completions"


def _tail_text(value: str, max_chars: int = 3000) -> str:
    text = (value or "").strip()
    if len(text) <= max_chars:
        return text
    return text[-max_chars:]


def _utc_now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _cleanup_export_daily_jobs_locked() -> None:
    if len(_export_daily_jobs) <= _MAX_EXPORT_DAILY_JOBS:
        return
    finished = sorted(
        (
            job
            for job in _export_daily_jobs.values()
            if job.get("status") in {"success", "failed"}
        ),
        key=lambda job: job.get("finished_at") or job.get("created_at") or "",
    )
    removable_count = len(_export_daily_jobs) - _MAX_EXPORT_DAILY_JOBS
    for job in finished[:removable_count]:
        _export_daily_jobs.pop(job["job_id"], None)


def _resolve_export_daily_script_path() -> Path:
    project_root = Path(__file__).resolve().parents[4]
    candidates = [
        Path(r"D:\Yida_project\Common\Web\Amazon\Flows\export_daily.py"),
        Path("/mnt/d/Yida_project/Common/Web/Amazon/Flows/export_daily.py"),
        project_root / "Common" / "Web" / "Amazon" / "Flows" / "export_daily.py",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise HTTPException(
        status_code=404,
        detail="抓取脚本不存在: D:\\Yida_project\\Common\\Web\\Amazon\\Flows\\export_daily.py",
    )


def _build_export_daily_urls(site: str) -> tuple[List[str], str]:
    rows = bsr_repo.fetch_latest_bsr_product_urls(site, limit=_EXPORT_DAILY_TOTAL_URLS)
    if not rows:
        raise HTTPException(status_code=404, detail=f"{site} 站点最新批次没有可抓取的 product_url")

    seen_urls: set[str] = set()
    payload_urls: List[str] = []
    batch_date = ""
    for row in rows:
        raw_url = str(row.get("product_url") or "").strip()
        if not raw_url:
            continue
        if not batch_date:
            value = row.get("createtime")
            batch_date = value.isoformat() if isinstance(value, date) else str(value or "")
        if raw_url in seen_urls:
            continue
        seen_urls.add(raw_url)
        payload_urls.append(raw_url)

    if not payload_urls:
        raise HTTPException(status_code=404, detail=f"{site} 站点最新批次 product_url 为空")
    if len(payload_urls) < _EXPORT_DAILY_TOTAL_URLS:
        raise HTTPException(
            status_code=400,
            detail=(
                f"{site} 站点最新批次仅有 {len(payload_urls)} 条有效 product_url，"
                f"需{_EXPORT_DAILY_TOTAL_URLS}条"
            ),
        )

    return payload_urls[:_EXPORT_DAILY_TOTAL_URLS], batch_date


def _resolve_export_daily_output_dir() -> Path:
    bi_amazon_root = Path(__file__).resolve().parents[3]
    target_dir = bi_amazon_root / "backend" / "files" / date.today().isoformat()
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir


def _chunk_urls(urls: List[str], batch_size: int) -> List[List[str]]:
    return [urls[i : i + batch_size] for i in range(0, len(urls), batch_size)]


def _extract_asin_from_text(text: str) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""
    match = re.search(r"/dp/([A-Z0-9]{10})", raw, flags=re.I)
    if match:
        return match.group(1).upper()
    match = re.search(r"/gp/product/([A-Z0-9]{10})", raw, flags=re.I)
    if match:
        return match.group(1).upper()
    match = re.search(r"\b([A-Z0-9]{10})\b", raw, flags=re.I)
    return match.group(1).upper() if match else ""


def _extract_asin_from_filename(path: Path) -> str:
    match = re.match(r"([A-Za-z0-9]{10})(?:_|$)", path.name)
    return match.group(1).upper() if match else ""


def _parse_daily_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: Any, *, default: Optional[Decimal] = None) -> Optional[Decimal]:
    if value is None or value == "":
        return default
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return default
    raw = str(value).strip()
    if not raw:
        return default
    raw = raw.replace(",", "").replace("$", "").replace("%", "")
    if not raw:
        return default
    try:
        return Decimal(raw)
    except InvalidOperation:
        return default


def _parse_int(value: Any, *, default: Optional[int] = None) -> Optional[int]:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    raw = str(value).strip().replace(",", "").replace("+", "")
    if not raw:
        return default
    try:
        return int(float(raw))
    except ValueError:
        return default


def _find_header_index(
    header_map: Dict[str, int],
    aliases: List[str],
    *,
    startswith: Optional[str] = None,
) -> Optional[int]:
    for alias in aliases:
        if alias in header_map:
            return header_map[alias]
    if startswith:
        for name, idx in header_map.items():
            if str(name).startswith(startswith):
                return idx
    return None


def _collect_export_workbooks(
    output_dir: Path,
    product_urls: List[str],
    started_at: Optional[datetime] = None,
) -> List[Path]:
    expected_asins = {_extract_asin_from_text(url) for url in product_urls}
    expected_asins.discard("")
    workbooks: List[Path] = []
    for path in sorted(output_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if started_at is not None:
            mtime = datetime.fromtimestamp(path.stat().st_mtime)
            if mtime < started_at:
                continue
        if not expected_asins:
            workbooks.append(path)
            continue
        asin_from_name = _extract_asin_from_filename(path)
        if asin_from_name and asin_from_name in expected_asins:
            workbooks.append(path)
    return workbooks


def _parse_fact_bsr_daily_rows_from_workbook(path: Path, site: str) -> List[tuple[Any, ...]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"缺少 openpyxl，无法解析导出文件: {exc}")

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if not wb.sheetnames:
            return []
        ws = wb[wb.sheetnames[0]]
        asin = str(ws.title or "").strip().upper() or _extract_asin_from_filename(path)
        if not asin:
            return []

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return []
        header_map = {str(col).strip(): idx for idx, col in enumerate(header_row) if col is not None}
        missing_headers = [col for col in _BSR_DAILY_REQUIRED_HEADERS if col not in header_map]
        if missing_headers:
            raise ValueError(f"{path.name} 缺少列: {', '.join(missing_headers)}")

        idx_date = header_map["日期"]
        idx_buybox = header_map["Buybox价格($)"]
        idx_price = header_map["价格($)"]
        idx_prime = header_map["Prime价格($)"]
        idx_coupon_price = header_map["Coupon价格($)"]
        idx_coupon_discount = header_map["Coupon折扣"]
        idx_child_sales = header_map["子体销量"]
        idx_fba_price = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["fba_price"])
        idx_fbm_price = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["fbm_price"])
        idx_strikethrough_price = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["strikethrough_price"])
        idx_bsr_rank = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["bsr_rank"])
        idx_bsr_reciprocating = _find_header_index(header_map, [], startswith="BSR[")
        idx_rating = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["rating"])
        idx_rating_count = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["rating_count"])
        idx_seller_count = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["seller_count"])

        rows: List[tuple[Any, ...]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            row_date = _parse_daily_date(row[idx_date] if idx_date < len(row) else None)
            if row_date is None:
                continue
            buybox_price = _parse_decimal(row[idx_buybox] if idx_buybox < len(row) else None, default=Decimal("0"))
            price = _parse_decimal(row[idx_price] if idx_price < len(row) else None, default=Decimal("0"))
            prime_price = _parse_decimal(row[idx_prime] if idx_prime < len(row) else None, default=None)
            coupon_price = _parse_decimal(row[idx_coupon_price] if idx_coupon_price < len(row) else None, default=None)
            coupon_discount = _parse_decimal(row[idx_coupon_discount] if idx_coupon_discount < len(row) else None, default=None)
            child_sales = _parse_int(row[idx_child_sales] if idx_child_sales < len(row) else None, default=None)
            fba_price = _parse_decimal(
                row[idx_fba_price] if idx_fba_price is not None and idx_fba_price < len(row) else None,
                default=None,
            )
            fbm_price = _parse_decimal(
                row[idx_fbm_price] if idx_fbm_price is not None and idx_fbm_price < len(row) else None,
                default=None,
            )
            strikethrough_price = _parse_decimal(
                row[idx_strikethrough_price] if idx_strikethrough_price is not None and idx_strikethrough_price < len(row) else None,
                default=None,
            )
            bsr_rank = _parse_int(
                row[idx_bsr_rank] if idx_bsr_rank is not None and idx_bsr_rank < len(row) else None,
                default=None,
            )
            bsr_reciprocating_saw_blades = _parse_int(
                row[idx_bsr_reciprocating] if idx_bsr_reciprocating is not None and idx_bsr_reciprocating < len(row) else None,
                default=None,
            )
            rating = _parse_decimal(
                row[idx_rating] if idx_rating is not None and idx_rating < len(row) else None,
                default=None,
            )
            rating_count = _parse_int(
                row[idx_rating_count] if idx_rating_count is not None and idx_rating_count < len(row) else None,
                default=None,
            )
            seller_count = _parse_int(
                row[idx_seller_count] if idx_seller_count is not None and idx_seller_count < len(row) else None,
                default=None,
            )
            rows.append(
                (
                    site,
                    asin,
                    row_date,
                    buybox_price,
                    price,
                    prime_price,
                    coupon_price,
                    coupon_discount,
                    child_sales,
                    fba_price,
                    fbm_price,
                    strikethrough_price,
                    bsr_rank,
                    bsr_reciprocating_saw_blades,
                    rating,
                    rating_count,
                    seller_count,
                )
            )
        return rows
    finally:
        wb.close()


def _import_fact_bsr_daily_from_exports(
    output_dir: Path,
    site: str,
    product_urls: List[str],
    started_at: Optional[datetime] = None,
) -> Dict[str, Any]:
    workbook_paths = _collect_export_workbooks(output_dir, product_urls, started_at)
    if not workbook_paths:
        raise HTTPException(status_code=500, detail=f"未找到可入库的导出文件（目录: {output_dir}）")

    all_rows: List[tuple[Any, ...]] = []
    failed_files: List[str] = []
    for path in workbook_paths:
        try:
            all_rows.extend(_parse_fact_bsr_daily_rows_from_workbook(path, site))
        except Exception as exc:
            failed_files.append(f"{path.name}: {exc}")

    if not all_rows:
        detail = "未解析到可入库数据"
        if failed_files:
            detail = f"{detail}；失败文件: {' | '.join(failed_files[:5])}"
        raise HTTPException(status_code=500, detail=detail)

    imported_rows = bsr_repo.upsert_fact_bsr_daily_rows(all_rows)
    return {
        "imported_rows": imported_rows,
        "workbook_count": len(workbook_paths),
        "failed_files": failed_files,
    }


def _execute_export_daily(site: Optional[str] = None) -> Dict[str, Any]:
    script_path = _resolve_export_daily_script_path()
    normalized_site = normalize_site(site)
    project_root = script_path.parents[4]
    product_urls, batch_date = _build_export_daily_urls(normalized_site)
    output_dir = _resolve_export_daily_output_dir()
    run_started_at = datetime.now()
    bootstrap_code = (
        "import importlib, sys; "
        "flow=importlib.import_module('Common.Web.Amazon.Flows.export_daily'); "
        "flow.CONFIG['export_dir']=sys.argv[1]; "
        "sys.argv=['export_daily', *sys.argv[2:]]; "
        "flow.main()"
    )
    env = os.environ.copy()
    batches = _chunk_urls(product_urls, _EXPORT_DAILY_BATCH_SIZE)
    stdout_logs: List[str] = []
    for idx, url_batch in enumerate(batches, start=1):
        cmd = [sys.executable, "-c", bootstrap_code, str(output_dir)]
        for url in url_batch:
            cmd.extend(["--url", url])
        try:
            completed = subprocess.run(
                cmd,
                cwd=str(project_root),
                env=env,
                capture_output=True,
                text=True,
                timeout=1800,
                check=False,
            )
        except subprocess.TimeoutExpired:
            raise HTTPException(status_code=504, detail=f"抓取脚本第{idx}/{len(batches)}批执行超时（30分钟）")
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"抓取脚本第{idx}/{len(batches)}批执行异常: {exc}")

        stdout_tail = _tail_text(completed.stdout or "")
        stderr_tail = _tail_text(completed.stderr or "")
        batch_log = f"[batch {idx}/{len(batches)} urls={len(url_batch)}]\n{stdout_tail}".strip()
        if batch_log:
            stdout_logs.append(batch_log)
        if completed.returncode != 0:
            detail = stderr_tail or stdout_tail or f"exit code {completed.returncode}"
            raise HTTPException(status_code=500, detail=f"抓取脚本第{idx}/{len(batches)}批执行失败: {detail}")

    daily_import = _import_fact_bsr_daily_from_exports(output_dir, normalized_site, product_urls, run_started_at)

    return {
        "message": "抓取完成",
        "site": normalized_site,
        "batch_date": batch_date,
        "url_count": len(product_urls),
        "batch_size": _EXPORT_DAILY_BATCH_SIZE,
        "batch_count": len(batches),
        "export_dir": str(output_dir),
        "script": str(script_path),
        "daily_rows": daily_import["imported_rows"],
        "daily_workbooks": daily_import["workbook_count"],
        "daily_failed_files": daily_import["failed_files"],
        "stdout_tail": _tail_text("\n\n".join(stdout_logs)),
    }


def _update_export_daily_job(job_id: str, **kwargs: Any) -> None:
    with _export_daily_jobs_lock:
        job = _export_daily_jobs.get(job_id)
        if not job:
            return
        job.update(kwargs)


def _run_export_daily_job(job_id: str, site: str) -> None:
    _update_export_daily_job(
        job_id,
        status="running",
        started_at=_utc_now_iso(),
        updated_at=_utc_now_iso(),
    )
    try:
        result = _execute_export_daily(site)
        _update_export_daily_job(
            job_id,
            status="success",
            finished_at=_utc_now_iso(),
            updated_at=_utc_now_iso(),
            result=result,
            error_message=None,
        )
    except HTTPException as exc:
        detail = exc.detail if isinstance(exc.detail, str) else str(exc.detail)
        _update_export_daily_job(
            job_id,
            status="failed",
            finished_at=_utc_now_iso(),
            updated_at=_utc_now_iso(),
            error_message=detail,
        )
    except Exception as exc:
        _update_export_daily_job(
            job_id,
            status="failed",
            finished_at=_utc_now_iso(),
            updated_at=_utc_now_iso(),
            error_message=str(exc),
        )


def submit_export_daily_job(site: Optional[str], operator_userid: str) -> Dict[str, Any]:
    normalized_site = normalize_site(site)
    job_id = uuid.uuid4().hex
    now = _utc_now_iso()
    job = {
        "job_id": job_id,
        "site": normalized_site,
        "status": "pending",
        "created_at": now,
        "updated_at": now,
        "started_at": None,
        "finished_at": None,
        "operator_userid": operator_userid,
        "error_message": None,
        "result": None,
    }
    with _export_daily_jobs_lock:
        _export_daily_jobs[job_id] = job
        _cleanup_export_daily_jobs_locked()

    worker = threading.Thread(
        target=_run_export_daily_job,
        args=(job_id, normalized_site),
        name=f"export-daily-{job_id[:8]}",
        daemon=True,
    )
    worker.start()
    return dict(job)


def get_export_daily_job(job_id: str, requester_userid: str, requester_role: str) -> Dict[str, Any]:
    with _export_daily_jobs_lock:
        job = _export_daily_jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="抓取任务不存在")
        if requester_role != "admin" and job.get("operator_userid") != requester_userid:
            raise HTTPException(status_code=403, detail="无权访问该抓取任务")
        return dict(job)


def run_export_daily(site: Optional[str] = None) -> Dict[str, Any]:
    return _execute_export_daily(site)


def to_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def to_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, Decimal):
        value = int(value)
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def price_to_string(value: Any) -> str:
    if value is None:
        return "$0.00"
    if isinstance(value, Decimal):
        value = float(value)
    try:
        return f"${float(value):.2f}"
    except (TypeError, ValueError):
        return "$0.00"


def parse_tags(value: Any) -> List[str]:
    if not value:
        return []
    return [tag.strip() for tag in str(value).split(",") if tag.strip()]


def split_asins(value: Optional[str]) -> List[str]:
    if not value:
        return []
    parts = re.split(r"[,，;|]", str(value))
    return [part.strip() for part in parts if part and part.strip()]


def unique_asins(values: List[str]) -> List[str]:
    seen: set[str] = set()
    result: List[str] = []
    for val in values:
        if val in seen:
            continue
        seen.add(val)
        result.append(val)
    return result


def derive_bsr_type(type_value: Any) -> str:
    if type_value is None:
        return "0"
    raw = str(type_value).strip()
    if raw == "1":
        return "1"
    if raw == "0":
        return "0"
    return "0"


def bsr_row_to_item(row: Dict[str, Any]) -> Dict[str, Any]:
    item_type = derive_bsr_type(row.get("type"))
    return {
        "rank": to_int(row.get("bsr_rank")),
        "bsr_rank": to_int(row.get("bsr_rank")),
        "site": row.get("site") or DEFAULT_BSR_SITE,
        "asin": row.get("asin") or "",
        "yida_asin": row.get("yida_asin") or "",
        "parent_asin": row.get("parent_asin") or "",
        "title": row.get("title") or "",
        "brand": row.get("brand") or "",
        "price": price_to_string(row.get("price")),
        "list_price": price_to_string(row.get("list_price")),
        "rating": to_float(row.get("score")),
        "score": to_float(row.get("score")),
        "reviews": to_int(row.get("comment_count")),
        "comment_count": to_int(row.get("comment_count")),
        "tags": parse_tags(row.get("tags")),
        "status": item_type,
        "type": item_type,
        "image_url": row.get("image_url") or "",
        "product_url": row.get("product_url") or "",
        "category_rank": to_int(row.get("category_rank")),
        "variation_count": to_int(row.get("variation_count")),
        "conversion_rate": to_float(row.get("conversion_rate")),
        "conversion_rate_period": row.get("conversion_rate_period"),
        "organic_traffic_count": to_int(row.get("organic_traffic_count")),
        "ad_traffic_count": to_int(row.get("ad_traffic_count")),
        "organic_search_terms": to_int(row.get("organic_search_terms")),
        "ad_search_terms": to_int(row.get("ad_search_terms")),
        "search_recommend_terms": to_int(row.get("search_recommend_terms")),
        "launch_date": row.get("launch_date").isoformat()
        if isinstance(row.get("launch_date"), date)
        else None,
        "sales_volume": to_int(row.get("sales_volume")),
        "sales": to_float(row.get("sales")),
        "createtime": row.get("createtime").isoformat()
        if isinstance(row.get("createtime"), date)
        else None,
    }


def list_bsr_items(limit: int, offset: int, createtime: Optional[date], site: str, role: str, userid: str) -> Dict[str, Any]:
    site = normalize_site(site)
    rows = bsr_repo.fetch_bsr_items(site, createtime, limit, offset, role, userid)
    items = []
    batch_date = None
    for row in rows:
        if batch_date is None and isinstance(row.get("createtime"), date):
            batch_date = row["createtime"].isoformat()
        items.append(bsr_row_to_item(row))
    return {"items": items, "batch_date": batch_date}


def lookup_bsr_item(
    asin: str,
    createtime: Optional[date],
    site: str,
    role: str,
    userid: str,
    brand: Optional[str] = None,
) -> Dict[str, Any]:
    site = normalize_site(site)
    normalized_brand = str(brand or "").strip() or None
    row = bsr_repo.fetch_bsr_lookup_row(asin, site, createtime, role, userid, normalized_brand)
    if not row:
        return {"item": None, "found": False}
    return {"item": bsr_row_to_item(row), "found": True}


def list_bsr_dates(site: str, limit: int, offset: int) -> List[str]:
    site = normalize_site(site)
    rows = bsr_repo.fetch_bsr_dates(site, limit, offset)
    items: List[str] = []
    for row in rows:
        value = row.get("createtime")
        items.append(value.isoformat() if isinstance(value, date) else str(value))
    return items


async def import_bsr_files(
    seller_file: Optional[UploadFile],
    seller_file_detail: Optional[UploadFile],
    jimu_file: Optional[UploadFile],
    jimu_file_51_100: Optional[UploadFile],
    site: str,
) -> Dict[str, Any]:
    normalized_site = normalize_site(site)
    has_detail = seller_file_detail is not None
    has_bundle = any([seller_file, jimu_file, jimu_file_51_100])
    if not has_detail and not has_bundle:
        raise HTTPException(status_code=400, detail="请上传卖家精灵明细（销量、销售额）或完整的明细数据三件套")
    if has_bundle and not (seller_file and jimu_file and jimu_file_51_100):
        raise HTTPException(
            status_code=400,
            detail="卖家精灵明细 + 极木与西柚#1-50 + 极木与西柚#51-100 必须一起上传",
        )

    tmp_dir = Path(tempfile.mkdtemp(prefix="bsr-import-"))
    try:
        saved: Dict[str, Path] = {}
        for key, upload in (
            ("seller_file", seller_file),
            ("seller_file_detail", seller_file_detail),
            ("jimu_file", jimu_file),
            ("jimu_file_51_100", jimu_file_51_100),
        ):
            if upload is None:
                continue
            filename = upload.filename or ""
            suffix = Path(filename).suffix.lower()
            if not suffix:
                raise HTTPException(status_code=400, detail="文件缺少扩展名")
            target = tmp_dir / f"{uuid.uuid4().hex}{suffix}"
            content = await upload.read()
            if not content:
                raise HTTPException(status_code=400, detail="上传文件为空")
            target.write_bytes(content)
            saved[key] = target

        seller_excel_path = saved.get("seller_file")
        seller_detail_path = saved.get("seller_file_detail")
        jimu_csv_path = saved.get("jimu_file")
        jimu_csv_next_path = saved.get("jimu_file_51_100")

        if has_bundle:
            if seller_excel_path is None or jimu_csv_path is None or jimu_csv_next_path is None:
                raise HTTPException(status_code=400, detail="明细导入缺少文件")
            if seller_excel_path.suffix.lower() not in {".xls", ".xlsx"}:
                raise HTTPException(status_code=400, detail="卖家精灵文件需为 Excel（.xls/.xlsx）")
            if jimu_csv_path.suffix.lower() != ".csv" or jimu_csv_next_path.suffix.lower() != ".csv":
                raise HTTPException(status_code=400, detail="极木与西柚文件需为 CSV（.csv）")

            from ..imports.bsr_importer import count_excel_rows, count_csv_rows

            seller_count = count_excel_rows(seller_excel_path)
            if seller_count != 100:
                raise HTTPException(
                    status_code=400,
                    detail=f"卖家精灵明细第一张表数据量应为100条，实际为{seller_count}条",
                )
            jimu_count = count_csv_rows(jimu_csv_path)
            if jimu_count != 50:
                raise HTTPException(
                    status_code=400,
                    detail=f"极木与西柚数据#1-50数据量应为50条，实际为{jimu_count}条",
                )
            jimu_next_count = count_csv_rows(jimu_csv_next_path)
            if jimu_next_count != 50:
                raise HTTPException(
                    status_code=400,
                    detail=f"极木与西柚数据#51-100数据量应为50条，实际为{jimu_next_count}条",
                )

        if has_detail:
            if seller_detail_path is None:
                raise HTTPException(status_code=400, detail="卖家精灵明细（销量、销售额）缺少文件")
            if seller_detail_path.suffix.lower() not in {".xls", ".xlsx"}:
                raise HTTPException(status_code=400, detail="卖家精灵明细需为 Excel（.xls/.xlsx）")
            from ..imports.bsr_importer import count_excel_rows

            seller_detail_count = count_excel_rows(seller_detail_path)
            if seller_detail_count != 100:
                raise HTTPException(
                    status_code=400,
                    detail=f"卖家精灵明细（销量、销售额）第一张表数据量应为100条，实际为{seller_detail_count}条",
                )

        insert_df = None
        monthly_df = None
        with get_connection() as conn:
            if has_bundle and seller_excel_path and jimu_csv_path and jimu_csv_next_path:
                bsr_repo.delete_bsr_items_for_today(normalized_site)
                insert_df = import_bsr_data(
                    str(seller_excel_path),
                    [str(jimu_csv_path), str(jimu_csv_next_path)],
                    connection=conn,
                    site=normalized_site,
                )
            if has_detail and seller_detail_path:
                monthly_df = import_bsr_monthly(
                    str(seller_detail_path),
                    connection=conn,
                    site=normalized_site,
                )

        return {
            "rows": len(insert_df) if insert_df is not None else 0,
            "monthly_rows": len(monthly_df) if monthly_df is not None else 0,
        }
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def list_bsr_monthly(asin: str, site: str, is_child: Optional[int] = None) -> List[Dict[str, Any]]:
    if not asin:
        raise HTTPException(status_code=400, detail="asin 不能为空")
    normalized_site = normalize_site(site)
    return bsr_repo.fetch_bsr_monthly(asin, normalized_site, is_child)


def list_bsr_daily(asin: str, site: str) -> List[Dict[str, Any]]:
    if not asin:
        raise HTTPException(status_code=400, detail="asin 不能为空")
    normalized_site = normalize_site(site)
    rows = bsr_repo.fetch_bsr_daily(asin, normalized_site)
    items: List[Dict[str, Any]] = []
    for row in rows:
        row_date = row.get("date")
        items.append(
            {
                "date": row_date.isoformat() if isinstance(row_date, date) else str(row_date or ""),
                "buybox_price": to_float(row.get("buybox_price"), 0.0),
                "price": to_float(row.get("price"), 0.0),
                "prime_price": to_float(row.get("prime_price"), 0.0),
                "coupon_price": to_float(row.get("coupon_price"), 0.0),
                "coupon_discount": to_float(row.get("coupon_discount"), 0.0),
                "child_sales": to_int(row.get("child_sales"), 0),
                "fba_price": to_float(row.get("fba_price"), 0.0),
                "fbm_price": to_float(row.get("fbm_price"), 0.0),
                "strikethrough_price": to_float(row.get("strikethrough_price"), 0.0),
                "bsr_rank": to_int(row.get("bsr_rank"), 0),
                "bsr_reciprocating_saw_blades": to_int(row.get("bsr_reciprocating_saw_blades"), 0),
                "rating": to_float(row.get("rating"), 0.0),
                "rating_count": to_int(row.get("rating_count"), 0),
                "seller_count": to_int(row.get("seller_count"), 0),
            }
        )
    return items


def _parse_openrouter_text(response_json: Dict[str, Any]) -> str:
    choices = response_json.get("choices")
    if not isinstance(choices, list):
        return ""
    for choice in choices:
        if not isinstance(choice, dict):
            continue
        message = choice.get("message")
        if not isinstance(message, dict):
            continue
        content = message.get("content")
        if isinstance(content, str) and content.strip():
            return content.strip()
        if not isinstance(content, list):
            continue
        fragments: List[str] = []
        for part in content:
            if not isinstance(part, dict):
                continue
            text = part.get("text")
            if isinstance(text, str) and text.strip():
                fragments.append(text.strip())
        if fragments:
            return "\n".join(fragments).strip()
    return ""


def _resolve_openrouter_model() -> str:
    explicit = str(os.getenv("OPENROUTER_MODEL", "")).strip() or str(os.getenv("OPEN_ROUTER_MODEL", "")).strip()
    if explicit:
        return explicit

    legacy = str(os.getenv("GEMINI_MODEL", "")).strip()
    if legacy:
        if legacy == "gemini-2.0-flash":
            return "google/gemini-2.0-flash-001"
        if "/" in legacy:
            return legacy
        return f"google/{legacy}"
    return "google/gemini-3-flash-preview"


def _build_openrouter_headers(api_key: str) -> Dict[str, str]:
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json; charset=utf-8",
    }
    referer = str(os.getenv("OPENROUTER_SITE_URL", "")).strip()
    if referer:
        headers["HTTP-Referer"] = referer
    app_name = str(os.getenv("OPENROUTER_APP_NAME", "Bi-Amazon Backend")).strip()
    if app_name:
        headers["X-Title"] = app_name
    return headers


def _build_bsr_ai_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    first = rows[0]
    last = rows[-1]
    first_price = to_float(first.get("price"), 0.0)
    last_price = to_float(last.get("price"), 0.0)
    first_child_sales = to_int(first.get("child_sales"), 0)
    last_child_sales = to_int(last.get("child_sales"), 0)
    coupon_days = sum(1 for row in rows if to_float(row.get("coupon_price"), 0.0) > 0)
    avg_price = round(sum(to_float(row.get("price"), 0.0) for row in rows) / max(len(rows), 1), 2)
    avg_buybox = round(sum(to_float(row.get("buybox_price"), 0.0) for row in rows) / max(len(rows), 1), 2)
    max_child_sales = max(to_int(row.get("child_sales"), 0) for row in rows)
    min_child_sales = min(to_int(row.get("child_sales"), 0) for row in rows)
    return {
        "date_start": str(first.get("date") or ""),
        "date_end": str(last.get("date") or ""),
        "first_price": first_price,
        "last_price": last_price,
        "first_child_sales": first_child_sales,
        "last_child_sales": last_child_sales,
        "coupon_days": coupon_days,
        "avg_price": avg_price,
        "avg_buybox_price": avg_buybox,
        "max_child_sales": max_child_sales,
        "min_child_sales": min_child_sales,
    }


def _call_openrouter_bsr_ai_insight(
    asin: str,
    site: str,
    range_days: int,
    rows: List[Dict[str, Any]],
    summary: Dict[str, Any],
) -> str:
    api_key = (
        str(os.getenv("OPENROUTER_API_KEY", "")).strip()
        or str(os.getenv("OPEN_ROUTER_API_KEY", "")).strip()
    )
    if not api_key:
        raise HTTPException(status_code=500, detail="OPENROUTER_API_KEY 未配置")

    model = _resolve_openrouter_model()

    rows_payload = []
    for row in rows:
        rows_payload.append(
            {
                "date": str(row.get("date") or ""),
                "buybox_price": to_float(row.get("buybox_price"), 0.0),
                "price": to_float(row.get("price"), 0.0),
                "prime_price": to_float(row.get("prime_price"), 0.0),
                "coupon_price": to_float(row.get("coupon_price"), 0.0),
                "coupon_discount": to_float(row.get("coupon_discount"), 0.0),
                "child_sales": to_int(row.get("child_sales"), 0),
                "fba_price": to_float(row.get("fba_price"), 0.0),
                "fbm_price": to_float(row.get("fbm_price"), 0.0),
                "strikethrough_price": to_float(row.get("strikethrough_price"), 0.0),
                "bsr_rank": to_int(row.get("bsr_rank"), 0),
                "bsr_reciprocating_saw_blades": to_int(row.get("bsr_reciprocating_saw_blades"), 0),
                "rating": to_float(row.get("rating"), 0.0),
                "rating_count": to_int(row.get("rating_count"), 0),
                "seller_count": to_int(row.get("seller_count"), 0),
            }
        )
    prompt = build_keepa_report_user_prompt(
        asin=asin,
        site=site,
        range_days=range_days,
        summary=summary,
        rows_payload=rows_payload,
    )
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": KEEPA_REPORT_SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.2,
        "top_p": 0.9,
        "max_tokens": 3072,
    }
    raw_payload = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        _OPENROUTER_API_BASE,
        data=raw_payload,
        method="POST",
        headers=_build_openrouter_headers(api_key),
    )
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            raw = resp.read().decode("utf-8", errors="ignore")
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"OpenRouter 调用失败: {exc.code} {_tail_text(body, 500)}")
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"OpenRouter 网络错误: {exc.reason}")
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"OpenRouter 调用异常: {exc}")

    try:
        response_json = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=502, detail="OpenRouter 返回解析失败")

    text = _parse_openrouter_text(response_json)
    if not text:
        raise HTTPException(status_code=502, detail="OpenRouter 未返回可用文本")
    return text


def _call_gemini_bsr_ai_insight(
    asin: str,
    site: str,
    range_days: int,
    rows: List[Dict[str, Any]],
    summary: Dict[str, Any],
) -> str:
    return _call_openrouter_bsr_ai_insight(asin, site, range_days, rows, summary)


def get_bsr_ai_insight(asin: str, site: str, range_days: int) -> Dict[str, Any]:
    target_asin = str(asin or "").strip().upper()
    if not target_asin:
        raise HTTPException(status_code=400, detail="asin 不能为空")
    normalized_site = normalize_site(site)
    safe_range_days = range_days if range_days in {7, 30, 90, 180} else 90
    rows = bsr_repo.fetch_bsr_daily_window(target_asin, normalized_site, safe_range_days)
    if not rows:
        raise HTTPException(status_code=404, detail="该 ASIN 在 fact_bsr_daily 暂无可分析数据")
    summary = _build_bsr_ai_summary(rows)
    report = _call_openrouter_bsr_ai_insight(target_asin, normalized_site, safe_range_days, rows, summary)
    return {
        "asin": target_asin,
        "site": normalized_site,
        "range_days": safe_range_days,
        "summary": summary,
        "report": report,
    }


def update_bsr_tags(asin: str, tag_list: List[str], createtime: Optional[date], site: str, role: str, userid: str, username: str) -> Dict[str, Any]:
    tag_string = ",".join([tag.strip() for tag in tag_list if tag and tag.strip()])
    target_site = normalize_site(site)
    target_date = createtime or bsr_repo.resolve_bsr_createtime(asin, target_site, None)
    if not target_date:
        raise HTTPException(status_code=404, detail="ASIN not found for update")

    affected, exists = bsr_repo.update_bsr_tags(asin, tag_string, target_site, target_date, role, userid)
    if role != "admin" and affected == 0 and not exists:
        raise HTTPException(status_code=403, detail="Forbidden")
    if not exists:
        raise HTTPException(status_code=404, detail="ASIN not found for update")

    detail = f"tags={tag_string}"
    if target_date:
        detail += f", createtime={target_date.isoformat()}"
    detail += f", site={target_site}"
    user_service.log_audit(
        module="bsr",
        action="update_tags",
        target_id=asin,
        operator_userid=userid,
        operator_name=username,
        detail=detail,
    )

    return {
        "asin": asin,
        "tags": [t for t in tag_string.split(",") if t],
        "updated": affected,
    }


def update_bsr_mapping(asin: str, yida_asin: Optional[str], createtime: Optional[date], site: str, userid: str, username: str) -> Dict[str, Any]:
    target_site = normalize_site(site)
    resolved_date = bsr_repo.resolve_bsr_createtime(asin, target_site, createtime)
    if not resolved_date:
        raise HTTPException(status_code=404, detail="BSR date not found for ASIN")

    requested_asins = unique_asins(split_asins(yida_asin)) if yida_asin else []
    if any(val.lower() == asin.lower() for val in requested_asins):
        raise HTTPException(status_code=400, detail="Cannot map to the same ASIN")

    bsr_repo.update_bsr_mapping(asin, requested_asins, target_site, resolved_date, userid)

    detail = f"yida_asin={','.join(requested_asins)}"
    detail += f", createtime={resolved_date.isoformat()}"
    detail += f", site={target_site}"
    user_service.log_audit(
        module="bsr",
        action="update_mapping",
        target_id=asin,
        operator_userid=userid,
        operator_name=username,
        detail=detail,
    )

    return {"asin": asin, "yida_asin": ",".join(requested_asins)}
