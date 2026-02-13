from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional
import os
import re
import subprocess
import sys
import uuid

from fastapi import HTTPException

from ..core.celery_app import celery_app
from ..core.config import normalize_site
from ..repositories import bsr_repo, export_daily_job_repo
from .bsr_common_service import _tail_text

_EXPORT_DAILY_TASK_NAME = "bi_amazon.export_daily.run"
_EXPORT_DAILY_TOTAL_URLS = 100
_EXPORT_DAILY_BATCH_SIZE = 10
_BSR_DAILY_REQUIRED_HEADERS = [
    "日期",
    "Buybox价格($)",
    "价格($)",
    "Prime价格($)",
    "Coupon价格($)",
    "Coupon折扣",
    "子体销量",
]

_BSR_DAILY_OPTIONAL_HEADERS = {
    "fba_price": ["FBA价格($)", "FBA价格", "FBA Price($)", "FBA Price"],
    "fbm_price": ["FBM价格($)", "FBM价格", "FBM Price($)", "FBM Price"],
    "strikethrough_price": ["划线价格($)", "划线价格", "List价格($)", "List Price($)"],
    "bsr_rank": ["BSR排名", "BSR Rank"],
    "rating": ["评分", "Rating"],
    "rating_count": ["评分数", "Rating Count", "Review Count"],
    "seller_count": ["卖家数", "Seller Count"],
}


def _resolve_export_daily_script_path() -> Path:
    project_root = Path(__file__).resolve().parents[4]
    candidates = [
        Path(r"D:\Yida_project\Common\Web\Amazon\Flows\export_daily.py"),
        Path("/mnt/d/Yida_project/Common/Web/Amazon/Flows/export_daily.py"),
        project_root / "Common" / "Web" / "Amazon" / "Flows" / "export_daily.py",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise HTTPException(
        status_code=404,
        detail="抓取脚本不存在: D:\\Yida_project\\Common\\Web\\Amazon\\Flows\\export_daily.py",
    )


def _build_export_daily_urls(site: str) -> tuple[List[str], str]:
    rows = bsr_repo.fetch_latest_bsr_product_urls(site, limit=_EXPORT_DAILY_TOTAL_URLS)
    if not rows:
        raise HTTPException(status_code=404, detail=f"{site} 站点最新批次没有可抓取的 product_url")

    seen_urls: set[str] = set()
    payload_urls: List[str] = []
    batch_date = ""
    for row in rows:
        raw_url = str(row.get("product_url") or "").strip()
        if not raw_url:
            continue
        if not batch_date:
            value = row.get("createtime")
            batch_date = value.isoformat() if isinstance(value, date) else str(value or "")
        if raw_url in seen_urls:
            continue
        seen_urls.add(raw_url)
        payload_urls.append(raw_url)

    if not payload_urls:
        raise HTTPException(status_code=404, detail=f"{site} 站点最新批次 product_url 为空")
    if len(payload_urls) < _EXPORT_DAILY_TOTAL_URLS:
        raise HTTPException(
            status_code=400,
            detail=(
                f"{site} 站点最新批次仅有 {len(payload_urls)} 条有效 product_url，"
                f"需{_EXPORT_DAILY_TOTAL_URLS}条"
            ),
        )

    return payload_urls[:_EXPORT_DAILY_TOTAL_URLS], batch_date


def _resolve_export_daily_output_dir() -> Path:
    bi_amazon_root = Path(__file__).resolve().parents[3]
    target_dir = bi_amazon_root / "backend" / "files" / date.today().isoformat()
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir


def _chunk_urls(urls: List[str], batch_size: int) -> List[List[str]]:
    return [urls[i : i + batch_size] for i in range(0, len(urls), batch_size)]


def _extract_asin_from_text(text: str) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""
    match = re.search(r"/dp/([A-Z0-9]{10})", raw, flags=re.I)
    if match:
        return match.group(1).upper()
    match = re.search(r"/gp/product/([A-Z0-9]{10})", raw, flags=re.I)
    if match:
        return match.group(1).upper()
    match = re.search(r"\b([A-Z0-9]{10})\b", raw, flags=re.I)
    return match.group(1).upper() if match else ""


def _extract_asin_from_filename(path: Path) -> str:
    match = re.match(r"([A-Za-z0-9]{10})(?:_|$)", path.name)
    return match.group(1).upper() if match else ""


def _parse_daily_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: Any, *, default: Optional[Decimal] = None) -> Optional[Decimal]:
    if value is None or value == "":
        return default
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return default
    raw = str(value).strip()
    if not raw:
        return default
    raw = raw.replace(",", "").replace("$", "").replace("%", "")
    if not raw:
        return default
    try:
        return Decimal(raw)
    except InvalidOperation:
        return default


def _parse_int(value: Any, *, default: Optional[int] = None) -> Optional[int]:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    raw = str(value).strip().replace(",", "").replace("+", "")
    if not raw:
        return default
    try:
        return int(float(raw))
    except ValueError:
        return default


def _find_header_index(
    header_map: Dict[str, int],
    aliases: List[str],
    *,
    startswith: Optional[str] = None,
) -> Optional[int]:
    for alias in aliases:
        if alias in header_map:
            return header_map[alias]
    if startswith:
        for name, idx in header_map.items():
            if str(name).startswith(startswith):
                return idx
    return None


def _collect_export_workbooks(
    output_dir: Path,
    product_urls: List[str],
    started_at: Optional[datetime] = None,
) -> List[Path]:
    expected_asins = {_extract_asin_from_text(url) for url in product_urls}
    expected_asins.discard("")
    workbooks: List[Path] = []
    for path in sorted(output_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if started_at is not None:
            mtime = datetime.fromtimestamp(path.stat().st_mtime)
            if mtime < started_at:
                continue
        if not expected_asins:
            workbooks.append(path)
            continue
        asin_from_name = _extract_asin_from_filename(path)
        if asin_from_name and asin_from_name in expected_asins:
            workbooks.append(path)
    return workbooks


def _parse_fact_bsr_daily_rows_from_workbook(path: Path, site: str) -> List[tuple[Any, ...]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"缺少 openpyxl，无法解析导出文件: {exc}")

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if not wb.sheetnames:
            return []
        ws = wb[wb.sheetnames[0]]
        asin = str(ws.title or "").strip().upper() or _extract_asin_from_filename(path)
        if not asin:
            return []

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return []
        header_map = {str(col).strip(): idx for idx, col in enumerate(header_row) if col is not None}
        missing_headers = [col for col in _BSR_DAILY_REQUIRED_HEADERS if col not in header_map]
        if missing_headers:
            raise ValueError(f"{path.name} 缺少列: {', '.join(missing_headers)}")

        idx_date = header_map["日期"]
        idx_buybox = header_map["Buybox价格($)"]
        idx_price = header_map["价格($)"]
        idx_prime = header_map["Prime价格($)"]
        idx_coupon_price = header_map["Coupon价格($)"]
        idx_coupon_discount = header_map["Coupon折扣"]
        idx_child_sales = header_map["子体销量"]
        idx_fba_price = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["fba_price"])
        idx_fbm_price = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["fbm_price"])
        idx_strikethrough_price = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["strikethrough_price"])
        idx_bsr_rank = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["bsr_rank"])
        idx_bsr_reciprocating = _find_header_index(header_map, [], startswith="BSR[")
        idx_rating = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["rating"])
        idx_rating_count = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["rating_count"])
        idx_seller_count = _find_header_index(header_map, _BSR_DAILY_OPTIONAL_HEADERS["seller_count"])

        rows: List[tuple[Any, ...]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            row_date = _parse_daily_date(row[idx_date] if idx_date < len(row) else None)
            if row_date is None:
                continue
            buybox_price = _parse_decimal(row[idx_buybox] if idx_buybox < len(row) else None, default=Decimal("0"))
            price = _parse_decimal(row[idx_price] if idx_price < len(row) else None, default=Decimal("0"))
            prime_price = _parse_decimal(row[idx_prime] if idx_prime < len(row) else None, default=None)
            coupon_price = _parse_decimal(row[idx_coupon_price] if idx_coupon_price < len(row) else None, default=None)
            coupon_discount = _parse_decimal(row[idx_coupon_discount] if idx_coupon_discount < len(row) else None, default=None)
            child_sales = _parse_int(row[idx_child_sales] if idx_child_sales < len(row) else None, default=None)
            fba_price = _parse_decimal(
                row[idx_fba_price] if idx_fba_price is not None and idx_fba_price < len(row) else None,
                default=None,
            )
            fbm_price = _parse_decimal(
                row[idx_fbm_price] if idx_fbm_price is not None and idx_fbm_price < len(row) else None,
                default=None,
            )
            strikethrough_price = _parse_decimal(
                row[idx_strikethrough_price] if idx_strikethrough_price is not None and idx_strikethrough_price < len(row) else None,
                default=None,
            )
            bsr_rank = _parse_int(
                row[idx_bsr_rank] if idx_bsr_rank is not None and idx_bsr_rank < len(row) else None,
                default=None,
            )
            bsr_reciprocating_saw_blades = _parse_int(
                row[idx_bsr_reciprocating] if idx_bsr_reciprocating is not None and idx_bsr_reciprocating < len(row) else None,
                default=None,
            )
            rating = _parse_decimal(
                row[idx_rating] if idx_rating is not None and idx_rating < len(row) else None,
                default=None,
            )
            rating_count = _parse_int(
                row[idx_rating_count] if idx_rating_count is not None and idx_rating_count < len(row) else None,
                default=None,
            )
            seller_count = _parse_int(
                row[idx_seller_count] if idx_seller_count is not None and idx_seller_count < len(row) else None,
                default=None,
            )
            rows.append(
                (
                    site,
                    asin,
                    row_date,
                    buybox_price,
                    price,
                    prime_price,
                    coupon_price,
                    coupon_discount,
                    child_sales,
                    fba_price,
                    fbm_price,
                    strikethrough_price,
                    bsr_rank,
                    bsr_reciprocating_saw_blades,
                    rating,
                    rating_count,
                    seller_count,
                )
            )
        return rows
    finally:
        wb.close()


def _import_fact_bsr_daily_from_exports(
    output_dir: Path,
    site: str,
    product_urls: List[str],
    started_at: Optional[datetime] = None,
) -> Dict[str, Any]:
    workbook_paths = _collect_export_workbooks(output_dir, product_urls, started_at)
    if not workbook_paths:
        raise HTTPException(status_code=500, detail=f"未找到可入库的导出文件（目录: {output_dir}）")

    all_rows: List[tuple[Any, ...]] = []
    failed_files: List[str] = []
    for path in workbook_paths:
        try:
            all_rows.extend(_parse_fact_bsr_daily_rows_from_workbook(path, site))
        except Exception as exc:
            failed_files.append(f"{path.name}: {exc}")

    if not all_rows:
        detail = "未解析到可入库数据"
        if failed_files:
            detail = f"{detail}；失败文件: {' | '.join(failed_files[:5])}"
        raise HTTPException(status_code=500, detail=detail)

    imported_rows = bsr_repo.upsert_fact_bsr_daily_rows(all_rows)
    return {
        "imported_rows": imported_rows,
        "workbook_count": len(workbook_paths),
        "failed_files": failed_files,
    }


def _execute_export_daily(site: Optional[str] = None) -> Dict[str, Any]:
    script_path = _resolve_export_daily_script_path()
    normalized_site = normalize_site(site)
    project_root = script_path.parents[4]
    product_urls, batch_date = _build_export_daily_urls(normalized_site)
    output_dir = _resolve_export_daily_output_dir()
    run_started_at = datetime.now()
    bootstrap_code = (
        "import importlib, sys; "
        "flow=importlib.import_module('Common.Web.Amazon.Flows.export_daily'); "
        "flow.CONFIG['export_dir']=sys.argv[1]; "
        "sys.argv=['export_daily', *sys.argv[2:]]; "
        "flow.main()"
    )
    env = os.environ.copy()
    batches = _chunk_urls(product_urls, _EXPORT_DAILY_BATCH_SIZE)
    stdout_logs: List[str] = []
    for idx, url_batch in enumerate(batches, start=1):
        cmd = [sys.executable, "-c", bootstrap_code, str(output_dir)]
        for url in url_batch:
            cmd.extend(["--url", url])
        try:
            completed = subprocess.run(
                cmd,
                cwd=str(project_root),
                env=env,
                capture_output=True,
                text=True,
                timeout=1800,
                check=False,
            )
        except subprocess.TimeoutExpired:
            raise HTTPException(status_code=504, detail=f"抓取脚本第{idx}/{len(batches)}批执行超时（30分钟）")
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"抓取脚本第{idx}/{len(batches)}批执行异常: {exc}")

        stdout_tail = _tail_text(completed.stdout or "")
        stderr_tail = _tail_text(completed.stderr or "")
        batch_log = f"[batch {idx}/{len(batches)} urls={len(url_batch)}]\n{stdout_tail}".strip()
        if batch_log:
            stdout_logs.append(batch_log)
        if completed.returncode != 0:
            detail = stderr_tail or stdout_tail or f"exit code {completed.returncode}"
            raise HTTPException(status_code=500, detail=f"抓取脚本第{idx}/{len(batches)}批执行失败: {detail}")

    daily_import = _import_fact_bsr_daily_from_exports(output_dir, normalized_site, product_urls, run_started_at)

    return {
        "message": "抓取完成",
        "site": normalized_site,
        "batch_date": batch_date,
        "url_count": len(product_urls),
        "batch_size": _EXPORT_DAILY_BATCH_SIZE,
        "batch_count": len(batches),
        "export_dir": str(output_dir),
        "script": str(script_path),
        "daily_rows": daily_import["imported_rows"],
        "daily_workbooks": daily_import["workbook_count"],
        "daily_failed_files": daily_import["failed_files"],
        "stdout_tail": _tail_text("\n\n".join(stdout_logs)),
    }


def run_export_daily_job(job_id: str, site: str) -> None:
    export_daily_job_repo.mark_job_running(job_id)
    try:
        result = _execute_export_daily(site)
        export_daily_job_repo.mark_job_success(job_id, result)
    except HTTPException as exc:
        detail = exc.detail if isinstance(exc.detail, str) else str(exc.detail)
        export_daily_job_repo.mark_job_failed(job_id, detail)
    except Exception as exc:
        export_daily_job_repo.mark_job_failed(job_id, str(exc))


def submit_export_daily_job(site: Optional[str], operator_userid: str) -> Dict[str, Any]:
    normalized_site = normalize_site(site)
    job_id = uuid.uuid4().hex
    export_daily_job_repo.insert_job(job_id, normalized_site, operator_userid)

    try:
        async_result = celery_app.send_task(_EXPORT_DAILY_TASK_NAME, args=[job_id, normalized_site])
        export_daily_job_repo.mark_job_queued(job_id, getattr(async_result, "id", None))
    except Exception as exc:
        export_daily_job_repo.mark_job_failed(job_id, f"任务入队失败: {exc}")
        raise HTTPException(status_code=502, detail=f"抓取任务入队失败: {exc}") from exc

    job = export_daily_job_repo.fetch_job(job_id)
    if not job:
        raise HTTPException(status_code=500, detail="抓取任务创建失败")
    return job


def get_export_daily_job(job_id: str, requester_userid: str, requester_role: str) -> Dict[str, Any]:
    job = export_daily_job_repo.fetch_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="抓取任务不存在")
    if requester_role != "admin" and job.get("operator_userid") != requester_userid:
        raise HTTPException(status_code=403, detail="无权访问该抓取任务")
    return job


def run_export_daily(site: Optional[str] = None) -> Dict[str, Any]:
    return _execute_export_daily(site)
